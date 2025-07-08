import sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
from datetime import datetime
from nepali_datetime import date as nepali_date
from PyQt5.QtWidgets import QFileDialog, QMessageBox

class ExcelHandler:
    HEADER_FILL = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    NOTE_FILL = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    
    @staticmethod
    def generate_template(db_path):
        """Generate enhanced Excel template with versioning and instructions"""
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        # Get column info, excluding auto-filled columns
        cursor.execute("PRAGMA table_info(member_info)")
        columns = [col[1] for col in cursor.fetchall() 
                  if col[1] not in ('id', 'date')]  # Skip auto-filled columns
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Member Import"
        
        # Add template metadata
        current_date = nepali_date.today().strftime('%Y-%m-%d %H:%M')
        ws.merge_cells('A1:D1')
        ws['A1'] = "MEMBER IMPORT TEMPLATE"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws['A2'] = f"Generated on: {current_date}"
        ws['A3'] = f"Template Version: 2.0"
        ws['A4'] = "NOTE: Date will be auto-filled with current BS date"
        
        # Format note area
        for row in ws.iter_rows(min_row=1, max_row=4, max_col=4):
            for cell in row:
                cell.fill = ExcelHandler.NOTE_FILL
        
        # Write headers starting from row 6
        header_row = 6
        for col_num, col_name in enumerate(columns, 1):
            cell = ws.cell(row=header_row, column=col_num, value=col_name.replace('_', ' ').title())
            cell.font = Font(bold=True)
            cell.fill = ExcelHandler.HEADER_FILL
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 22
            
            # Add data type hints
            type_cell = ws.cell(row=header_row+1, column=col_num, 
                              value=f"Type: {ExcelHandler._get_column_type(col_name)}")
            type_cell.font = Font(italic=True, color="808080")
        
        # Add example data
        example_data = {
            'member_number': '001000001',
            'member_name': 'John Doe',
            'phone': '9841234567',
            'dob_bs': '2050-01-15'
        }
        for col_num, col_name in enumerate(columns, 1):
            if col_name in example_data:
                example_cell = ws.cell(row=header_row+2, column=col_num, 
                                     value=example_data[col_name])
                example_cell.font = Font(color="00FF00")
                if 'date' in col_name or 'dob' in col_name:
                    example_cell.number_format = 'YYYY-MM-DD'
        
        # Add comprehensive instructions
        instructions = [
            ("INSTRUCTIONS:", "A10"),
            ("1. Fill data below the header row (Row 7)", "A11"),
            ("2. Do not modify column headers", "A12"),
            ("3. Date formats must be YYYY-MM-DD", "A13"),
            ("4. Member numbers must be unique", "A14"),
            ("5. Required fields: member_number, member_name", "A15")
        ]
        
        for text, cell_ref in instructions:
            ws[cell_ref] = text
            if "INSTRUCTIONS" in text:
                ws[cell_ref].font = Font(bold=True, color="FF0000")
        
        # Freeze header row
        ws.freeze_panes = "A7"
        
        conn.close()
        return wb
    
    @staticmethod
    def _get_column_type(col_name):
        """Enhanced column type mapping with validation rules"""
        type_map = {
            'member_number': 'Text (Unique, Required)',
            'member_name': 'Text (Required)',
            'phone': 'Number (10 digits)',
            'dob_bs': 'B.S. Date (YYYY-MM-DD)',
            'citizenship_no': 'Text/Number',
            'ward_no': 'Number (1-32)'
        }
        return type_map.get(col_name, 'Text')

    @staticmethod
    def import_data(db_path, filepath):
        """Enhanced import with automatic date, validation, and logging"""
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        try:
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
            
            # Find headers (skip metadata rows)
            # headers = []

            # for row in ws.iter_rows(values_only=True):
            #     if row and any("member_number" in str(cell).lower() for cell in row):
            #         # headers = [str(cell).lower().replace(' ', '_') for cell in row]
            #         headers = [str(cell).strip().lower().replace(' ', '_') for cell in row]
            #         data_start_row = ws.min_row + list(ws.iter_rows()).index(row) + 2
            #         break
            # Explicitly set header row index (row 6 in template)
            header_row_idx = 6
            header_row = [cell.value for cell in ws[header_row_idx]]

            # Clean header names
            headers = [str(cell).strip().lower().replace(' ', '_') for cell in header_row if cell]

            if not headers:
                return False, "Could not find valid column headers"
            
            data_start_row = header_row_idx + 1
            
            # Validate against database schema
            cursor.execute("PRAGMA table_info(member_info)")
            db_columns = [col[1] for col in cursor.fetchall() 
                         if col[1] not in ('id', 'date')]
            
            missing_cols = set(db_columns) - set(headers)
            if missing_cols:
                return False, f"Missing columns: {', '.join(missing_cols)}"
            
            # Get current BS date for all records
            current_date = nepali_date.today().strftime('%Y-%m-%d')
            
            # Prepare SQL with date included
            sql_columns = ['date'] + headers
            placeholders = ', '.join(['?'] * len(sql_columns))
            sql = f"INSERT INTO member_info ({', '.join(sql_columns)}) VALUES ({placeholders})"
            
            # Import statistics
            imported = 0
            skipped = 0
            errors = []
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True), 1):
                if not any(row):  # Skip empty rows
                    continue
                
                # Validate required fields
                if not row[headers.index('member_number')] or not row[headers.index('member_name')]:
                    errors.append(f"Row {row_idx}: Missing required fields")
                    skipped += 1
                    continue
                
                # Add current date and import
                try:
                    # cursor.execute(sql, [current_date] + list(row)) -> old code
                    cleaned_row = [str(cell).strip() if isinstance(cell, str) else cell for cell in row]
                    cursor.execute(sql, [current_date] + cleaned_row)

                    imported += 1
                except sqlite3.IntegrityError as e:
                    if "UNIQUE" in str(e):
                        errors.append(f"Row {row_idx}: Duplicate member number")
                    else:
                        errors.append(f"Row {row_idx}: {str(e)}")
                    skipped += 1
                except Exception as e:
                    errors.append(f"Row {row_idx}: {str(e)}")
                    skipped += 1
            
            conn.commit()
            
            # Prepare result message
            message = [
                f"Import completed on {current_date}",
                f"Successfully imported: {imported} members",
                f"Skipped: {skipped} rows"
            ]
            
            if errors:
                error_log_path = Path("logs/import_errors.log")
                error_log_path.parent.mkdir(exist_ok=True)
                with open(error_log_path, "a") as f:
                    f.write(f"\n\nImport on {datetime.now()}\n")
                    f.write("\n".join(errors))
                
                message.append(f"Errors logged to: {error_log_path}")
                message.append("First error: " + errors[0])
            
            return True, "\n".join(message)
            
        except Exception as e:
            conn.rollback()
            return False, f"Fatal import error: {str(e)}"
        finally:
            conn.close()